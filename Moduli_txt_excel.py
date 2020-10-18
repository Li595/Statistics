import os
import re
import openpyxl
from os.path import join, abspath
# Запрос к пользователю
filename = input("Введите путь к файлу: ")
data_path = join('.',filename)
data_path = abspath(data_path)
vyborka1 = []
vyborka2 = []
k = []
# Определение ресурса
extension = data_path.split('.')[-1]
if extension == 'txt':

    # Считывание данных из txt в список c проверкой значений
    with open(data_path, 'r+') as file:
        l = file.read().replace(',', '.').split()                        #       Замена запятых на точки для возможности работы с float
        count = 0
        for i in l:
          if '.' in i:                                                   #                         Проверка на наличие точки в значении
            try:                                                         #                и при наличии запись значения в единый список
                k.append(float(i))                                       #                            с проверкой на вещественное число
            except ValueError:
               print("Одно из значений не подошло под формат выборки")

        #   Деление списка на 2 выборки, где под чётными индексами 
        #   находятся элементы vyborka1, а под нечётными vyborka 2
        for i in k:
            if count % 2 == 0:
                vyborka1.append(float(i))
                count = count + 1
            else:
                vyborka2.append(float(i))
                count = count + 1
    
else : 

    #  Считывание данных из Excel с проверкой значений
    if extension == 'xlsx':
        wb = openpyxl.reader.excel.load_workbook(filename = data_path, data_only = True)         
        ws = wb.active                                                   #                 Присваиваем значения только активного листа

        #  Запись из строк, начиная с 3-й из разных столбцов в разные выборки
        for row in ws.iter_rows(min_row=3, min_col=1, max_row=ws.max_row, max_col=ws.max_column ):
            if len(row) > 0:
                    vyborka1.append(row[1].value)
                    vyborka2.append(row[2].value)
    else :
         print("false")
print(vyborka1)
print(vyborka2)




